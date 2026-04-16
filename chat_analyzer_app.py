"""
Chat Enquiries Daily Analyzer — Shopee & Lazada
────────────────────────────────────────────────
Run with:  streamlit run chat_analyzer_app.py

Install first:  pip install streamlit pandas openpyxl
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import re, io, warnings
warnings.filterwarnings('ignore')

st.set_page_config(page_title="Chat Analyzer — Shopee & Lazada", page_icon="🛒", layout="wide")

# ── HEADER ─────────────────────────────────────────────────────────────────────
st.title("🛒 Chat Enquiries Daily Analyzer")
st.caption("Shopee + Lazada  ·  CSAT  ·  Sentiment  ·  Unreplied / Unresolved  ·  Action Items per Account")

# ── KEYWORDS ────────────────────────────────────────────────────────────────────
AUTO_REPLY_PATTERNS = [
    # Welcome / greeting bots
    r'ยินดีต้อนรับ', r'แอดมินยินดีให้บริการ', r'ยินดีให้บริการ',
    r'สวัสดีค่ะ แอดมิน', r'สวัสดีครับ แอดมิน',
    r'kami sudah menerima', r'mohon ditunggu',
    r'thank you for contacting', r'we will get back to you',
    r'our team will respond', r"we'?ll reply",
    r'auto.?reply', r'automated',
    # Office hours / agent unavailable
    r'เวลาทำการ', r'นอกเวลาทำการ', r'ทีมงานจะกลับมา', r'พักกลางวัน', r'หยุดทำการ',
    r'outside.*hour', r'outside.*office',
    r'business hours', r'office hours',
    r'agent.*not available', r'agents.*not available',
    r'no agent', r'unavailable.*agent',
    r'\b9\s*[-–to]+\s*[56]\s*(pm|am)?\b',
    r'\b(8|9|10)\s*am\s*[-–to]+\s*(5|6|7)\s*pm\b',
    r'จ[-–]ศ\s*\d', r'จันทร์.*ศุกร์', r'monday.*friday.*\d',
    r'waktu (operasional|layanan)', r'jam operasional', r'jam kerja', r'di luar jam',
    r'outside.*working.*hour', r'currently.*unavailable',
    r'จะติดต่อกลับ.*โดยเร็ว', r'กลับมาให้บริการ.*โดยเร็ว',
]

STALLING_PATTERNS = [
    r'จะตรวจสอบ', r'กำลังตรวจสอบ', r'จะแจ้งกลับ', r'แจ้งกลับภายหลัง',
    r'ติดตามให้', r'กำลังดำเนินการ', r'รอสักครู่', r'รอก่อนนะ',
    r'จะรีบดำเนินการ', r'ขอเวลาตรวจสอบ', r'ขอตรวจสอบก่อน',
    r'จะประสานงาน', r'ประสานงานให้', r'แอดมินจะ.*ตรวจสอบ',
    r'ดำเนินการให้', r'ติดต่อกลับ', r'จะติดต่อกลับ',
    r"we'?ll? (check|look|get back|follow up|investigate)",
    r'let me (check|look into|verify|confirm)',
    r'(checking|looking into|investigating|following up)',
    r'will (check|get back|follow up|look into|update)',
    r'please (wait|hold on|allow us)',
    r'i will (check|get back|follow up)',
    r'we are (checking|looking|investigating)',
    r'we will (check|look|get back|follow up|update you)',
    r'kindly (wait|allow)', r'bear with us', r'get back to you',
    r'akan kami (cek|periksa|tindak lanjut)', r'kami sedang (cek|periksa|proses)',
    r'mohon (tunggu|ditunggu)', r'akan segera', r'kami akan (hubungi|informasikan|balas)',
    r'susuriin (namin|po)', r'magpapadala (kami|po)', r'aabisuhan (kita|po)',
]
ANGRY_KEYWORDS = [
    'ผิดหวัง','โกรธ','รำคาญ','ไม่พอใจ','แย่มาก','แย่','หลอกลวง','ห่วย','ยกเลิก',
    'คืนเงิน','เสียหาย','ไม่ได้รับ','ของเสีย','ของปลอม','ช้ามาก','รอนาน','ไม่ส่ง',
    'สินค้าไม่ตรง','ผิดสินค้า','หัก','แตก','เสีย','ชำรุด','ไม่ได้รับของ','หายไป',
    'terrible','worst','angry','disappointed','frustrated','cheated','scam','fraud',
    'fake','broken','damaged','wrong item','missing','never received','unacceptable',
    'horrible','awful','refund','cancel','complain','complaint','pathetic','useless',
    'tipu','bohong','rusak','cacat','mengecewakan','marah','kecewa','buruk',
    'jelek','parah','tidak diterima','hilang','salah barang','minta refund','batal',
    'sira','peke','hindi natanggap','basura',
]
HAPPY_KEYWORDS = [
    'ขอบคุณ','ขอบใจ','ดีมาก','ประทับใจ','พอใจ','สุดยอด','เยี่ยม','ขอบพระคุณ',
    'ดีเลย','ได้เลย','ขอบคุณค่ะ','ขอบคุณครับ','รับทราบ','ยอดเยี่ยม',
    'thank you','thanks','great','excellent','awesome','perfect','love it',
    'good','nice','happy','satisfied','wonderful','amazing','fantastic','superb',
    'appreciate','helpful','fast','quick','well done','recommend',
    'terima kasih','bagus','mantap','oke','baik','sip','keren','memuaskan','puas',
    'salamat','maganda','ayos','galing','masaya',
]
ISSUE_KEYWORDS = {
    'Shipping/Delivery':   ['จัดส่ง','ส่ง','delivery','shipping','ยังไม่ได้รับ','ไม่ส่ง','ช้า','ติดตาม','track','pengiriman','kurir'],
    'Wrong/Damaged Item':  ['ผิดสินค้า','สินค้าไม่ตรง','ของแตก','ของเสีย','wrong item','damaged','broken','defective','rusak','cacat'],
    'Refund/Cancel':       ['คืนเงิน','ยกเลิก','refund','cancel','return','batal','pengembalian'],
    'Stock/Availability':  ['หมด','out of stock','สินค้าหมด','stock','habis','tidak tersedia'],
    'Product Inquiry':     ['ราคา','price','สินค้า','product','item','สี','size','ขนาด','spec','รุ่น','model'],
    'Order Status':        ['สถานะ','order','คำสั่งซื้อ','status','pesanan','ออเดอร์'],
    'Payment':             ['ชำระ','payment','เงิน','โอน','pay','จ่าย','pembayaran','bayar'],
}
ACTION_MAP = {
    'Shipping/Delivery':  "1. Pull tracking in Seller Center  2. Check logistics status  3. Send buyer live tracking link  4. If >3 days delayed, escalate",
    'Wrong/Damaged Item': "1. Ask buyer for clear photos  2. Verify order in Seller Center  3. Initiate return/replacement  4. Escalate with photo evidence",
    'Refund/Cancel':      "1. Locate order in Seller Center  2. Process cancellation/refund  3. Confirm to buyer with timeline",
    'Stock/Availability': "1. Check warehouse stock  2. Update listing if OOS  3. Inform buyer or suggest alternatives",
    'Product Inquiry':    "1. Pull product specs  2. Gather photos/variants  3. Reply with complete details",
    'Order Status':       "1. Check order in Seller Center  2. Identify step  3. Share status + estimated delivery date",
    'Payment':            "1. Verify payment in Seller Center  2. Confirm processing  3. Send estimated dispatch date",
    'General Inquiry':    "1. Read full conversation  2. Identify core question  3. Give specific accurate answer",
}

def build_buyer_summary(buyer_texts, last_buyer_msg):
    """Concise 1-2 sentence summary of what the buyer said — like Gmail's email preview."""
    texts = [str(t).strip() for t in buyer_texts if pd.notna(t) and str(t).strip()]
    if not texts:
        return ''
    recent = texts[-3:]
    combined = ' | '.join(recent)
    if len(combined) > 220:
        combined = combined[:217] + '…'
    return combined

def build_contextual_reply(sentiment, issue, last_buyer_msg, buyer_texts):
    openers = {
        'ANGRY':   "We sincerely apologize for the difficulty you experienced.",
        'NEUTRAL': "Thank you for reaching out to us.",
        'HAPPY':   "Thank you for your kind message!",
    }
    resolutions = {
        'Shipping/Delivery':
            "Regarding your delivery concern: we are urgently checking your shipment status right now. "
            "We will send you the live tracking link and an exact delivery timeline within a few hours. "
            "If there is a delay we will escalate this immediately.",
        'Wrong/Damaged Item':
            "We are sorry to hear this. Could you please send us clear photos of the item you received? "
            "Once confirmed, we will arrange an immediate replacement or a full refund — whichever you prefer.",
        'Refund/Cancel':
            "We have noted your request and are processing it now. "
            "You will receive a confirmation once done. The refund will reflect within 3–5 business days.",
        'Stock/Availability':
            "We apologize — this item is currently out of stock. "
            "Please follow our store to be notified automatically when it is restocked.",
        'Product Inquiry':
            "Here are the details for the product you asked about: "
            "[fill in: size / material / colour options / specs]. "
            "Feel free to ask if you need more photos or have other questions.",
        'Order Status':
            "We have just checked your order. "
            "Current status: [fill in: processing / packed / handed to courier]. "
            "Estimated delivery: [fill in date]. We will keep you updated.",
        'Payment':
            "We have confirmed your payment. Your order is being processed and will be dispatched within "
            "[fill in: 1–2 business days]. We will send you the tracking number once shipped.",
        'General Inquiry':
            "Thank you for your message. "
            "[Please fill in the specific answer to the customer's question]. "
            "If you need any further help, we are happy to assist.",
    }
    customer_excerpt = str(last_buyer_msg).strip()[:150] if last_buyer_msg else ''
    if len(customer_excerpt) > 140: customer_excerpt += '…'
    resolution = resolutions.get(issue, resolutions['General Inquiry'])
    reply = f"{openers[sentiment]} {resolution}"
    if customer_excerpt and len(customer_excerpt) > 10:
        reply += f'\n\n[Customer said: "{customer_excerpt}"]'
    return reply

def is_auto_reply(msg):
    if pd.isna(msg): return False
    return any(re.search(p, str(msg).lower()) for p in AUTO_REPLY_PATTERNS)

def is_stalling(msg):
    if pd.isna(msg): return False
    return any(re.search(p, str(msg).lower()) for p in STALLING_PATTERNS)

def get_sentiment(messages):
    text = ' '.join([str(m).lower() for m in messages if pd.notna(m)])
    angry = sum(1 for kw in ANGRY_KEYWORDS if kw in text)
    happy = sum(1 for kw in HAPPY_KEYWORDS if kw in text)
    if angry > happy and angry >= 1: return 'ANGRY'
    if happy > angry and happy >= 1: return 'HAPPY'
    return 'NEUTRAL'

def get_issue(messages):
    text = ' '.join([str(m).lower() for m in messages if pd.notna(m)])
    for cat, kws in ISSUE_KEYWORDS.items():
        if any(kw in text for kw in kws): return cat
    return 'General Inquiry'

def csat(s): return {'HAPPY':5,'NEUTRAL':3,'ANGRY':1}[s]

def analyse_conv(grp):
    grp    = grp.sort_values('MESSAGE_TIME')
    buyer  = grp[grp['SENDER']=='buyer']
    seller = grp[grp['SENDER']=='seller']
    real_s = seller[~seller['IS_AUTO_REPLY']]
    last_real_seller = real_s['MESSAGE_PARSED'].iloc[-1] if len(real_s)>0 else None
    seller_stalled   = is_stalling(last_real_seller)
    last_sender      = grp.iloc[-1]['SENDER']
    only_auto        = len(seller)>0 and len(real_s)==0
    is_unreplied     = (last_sender=='buyer') or only_auto
    is_unresolved    = (grp['IS_ANSWERED'].iloc[-1]==False) or seller_stalled
    buyer_texts      = [m for m in buyer['MESSAGE_PARSED'].tolist() if pd.notna(m) and str(m).strip()]
    sentiment        = get_sentiment(buyer_texts)
    issue            = get_issue(buyer_texts + seller['MESSAGE_PARSED'].tolist())
    valid_buyer_msgs = buyer['MESSAGE_PARSED'].dropna()
    valid_buyer_msgs = valid_buyer_msgs[valid_buyer_msgs.str.strip() != '']
    last_buyer       = valid_buyer_msgs.iloc[-1] if len(valid_buyer_msgs) > 0 else ''
    buyer_summary    = build_buyer_summary(buyer_texts, last_buyer)
    reply  = build_contextual_reply(sentiment, issue, last_buyer, buyer_texts)
    action = ACTION_MAP.get(issue, ACTION_MAP['General Inquiry'])
    site_ids = grp['SITE_NICK_NAME_ID'].dropna().unique()
    return pd.Series({
        'PLATFORM':           grp['PLATFORM'].iloc[0],
        'STORE_CODE':         grp['STORE_CODE'].iloc[0],
        'SITE_NICK_NAME_ID':  site_ids[0] if len(site_ids)>0 else '',
        'COUNTRY_CODE':       grp['COUNTRY_CODE'].iloc[0],
        'BUYER_NAME':         grp['BUYER_NAME'].dropna().iloc[0] if grp['BUYER_NAME'].dropna().shape[0]>0 else 'Unknown',
        'FIRST_MSG_TIME':     grp['MESSAGE_TIME'].min(),
        'LAST_MSG_TIME':      grp['MESSAGE_TIME'].max(),
        'AGE_HOURS':          round((grp['MESSAGE_TIME'].max()-grp['MESSAGE_TIME'].min()).total_seconds()/3600,1),
        'TOTAL_MESSAGES':     len(grp),
        'BUYER_MESSAGES':     len(buyer),
        'SELLER_MESSAGES':    len(seller),
        'REAL_SELLER_REPLIES':len(real_s),
        'IS_UNREPLIED':       is_unreplied,
        'IS_UNRESOLVED':      is_unresolved,
        'ONLY_AUTO_REPLIED':  only_auto,
        'SELLER_STALLED':     seller_stalled,
        'SENTIMENT':          sentiment,
        'CSAT_SCORE':         csat(sentiment),
        'ISSUE_CATEGORY':     issue,
        'BUYER_CHAT_SUMMARY':  buyer_summary,
        'LAST_BUYER_MESSAGE':  str(last_buyer)[:300] if last_buyer else '',
        'SUGGESTED_REPLY':     reply,
        'ACTION_ITEM':         action,
    })
def store_summary(grp):
    unreplied  = int(grp['IS_UNREPLIED'].sum()) if 'IS_UNREPLIED' in grp.columns else 0
    unresolved = int(grp['IS_UNRESOLVED'].sum()) if 'IS_UNRESOLVED' in grp.columns else 0
    angry      = int((grp['SENTIMENT']=='ANGRY').sum()) if 'SENTIMENT' in grp.columns else 0

    # ✅ SAFE ISSUE
    if 'ISSUE_CATEGORY' in grp.columns:
if 'SENTIMENT' in grp.columns and 'BUYER_NAME' in grp.columns:
    angry_names = [
        str(x) for x in grp.loc[grp['SENTIMENT']=='ANGRY', 'BUYER_NAME']
        .dropna().unique()[:5]
    ]
else:
    angry_names = []
else:
        top_issues = []

    # ✅ SAFE NAMES
    if 'SENTIMENT' in grp.columns and 'BUYER_NAME' in grp.columns:
        angry_names = [
            str(x) for x in grp.loc[grp['SENTIMENT']=='ANGRY', 'BUYER_NAME']
            .dropna().unique()[:5]
        ]
    else:
        angry_names = []

    avg_csat = round(grp['CSAT_SCORE'].mean(), 2) if 'CSAT_SCORE' in grp.columns else 0
    country  = grp['COUNTRY_CODE'].mode().iloc[0] if 'COUNTRY_CODE' in grp.columns and len(grp)>0 else ''

    if   angry>=3 or unreplied>=10 or unresolved>=15: priority='P1 - CRITICAL'
    elif angry>=1 or unreplied>=5  or unresolved>=8:  priority='P2 - HIGH'
    elif unreplied>=2 or unresolved>=3:               priority='P3 - MEDIUM'
    else:                                             priority='P4 - LOW'

    actions = []
    if unreplied  > 0: actions.append(f"Reply to {unreplied} unreplied chat(s) immediately")
    if angry      > 0: actions.append(f"Address {angry} angry buyer(s): {', '.join(angry_names[:3])}")
    if unresolved > 0: actions.append(f"Resolve {unresolved} open ticket(s)")
    if avg_csat < 2.5: actions.append("CSAT critical — review all negative chats & escalate")
    if 'Shipping/Delivery'  in top_issues: actions.append("Investigate delivery delays")
    if 'Wrong/Damaged Item' in top_issues: actions.append("Review packing quality")
    if 'Refund/Cancel'      in top_issues: actions.append("Process pending refunds")

    platform, store_code, site_nick = grp.name

    return pd.Series({
        'PRIORITY': priority,
        'PLATFORM': platform,
        'STORE_CODE': store_code,
        'SITE_NICK_NAME_ID': site_nick,
        'COUNTRY': country,
        'TOTAL_CHATS_7D': len(grp),
        'UNREPLIED': unreplied,
        'AUTO_REPLY_ONLY': int(grp['ONLY_AUTO_REPLIED'].sum()) if 'ONLY_AUTO_REPLIED' in grp.columns else 0,
        'UNRESOLVED': unresolved,
        'ANGRY_BUYERS': angry,
        'NEUTRAL_BUYERS': int((grp['SENTIMENT']=='NEUTRAL').sum()) if 'SENTIMENT' in grp.columns else 0,
        'HAPPY_BUYERS': int((grp['SENTIMENT']=='HAPPY').sum()) if 'SENTIMENT' in grp.columns else 0,
        'AVG_CSAT': avg_csat,
        'TOP_ISSUES': ' | '.join(top_issues) if top_issues else '-',
        'ANGRY_BUYER_NAMES': ', '.join(angry_names) if angry_names else '-',
        'ACTION_ITEMS': '\n'.join(f"• {a}" for a in actions) if actions else '✓ No critical actions',
    })
    unreplied  = int(grp['IS_UNREPLIED'].sum())
    unresolved = int(grp['IS_UNRESOLVED'].sum())
    angry      = int((grp['SENTIMENT']=='ANGRY').sum())
if 'ISSUE_CATEGORY' in grp.columns:
    top_issues = grp['ISSUE_CATEGORY'].value_counts().head(3).index.tolist()
else:
    top_issues = []
    if 'SENTIMENT' in grp.columns and 'BUYER_NAME' in grp.columns:
    angry_names = [
        str(x) for x in grp.loc[grp['SENTIMENT']=='ANGRY', 'BUYER_NAME']
        .dropna().unique()[:5]
    ]
else:
    angry_names = []
    avg_csat   = round(grp['CSAT_SCORE'].mean(), 2)
    country    = grp['COUNTRY_CODE'].mode().iloc[0] if len(grp)>0 else ''
    if   angry>=3 or unreplied>=10 or unresolved>=15: priority='P1 - CRITICAL'
    elif angry>=1 or unreplied>=5  or unresolved>=8:  priority='P2 - HIGH'
    elif unreplied>=2 or unresolved>=3:               priority='P3 - MEDIUM'
    else:                                             priority='P4 - LOW'
    actions = []
    if unreplied  > 0: actions.append(f"Reply to {unreplied} unreplied chat(s) immediately")
    if angry      > 0: actions.append(f"Address {angry} angry buyer(s): {', '.join(angry_names[:3])}")
    if unresolved > 0: actions.append(f"Resolve {unresolved} open ticket(s)")
    if avg_csat < 2.5: actions.append("CSAT critical — review all negative chats & escalate")
    if 'Shipping/Delivery'  in top_issues: actions.append("Investigate delivery delays")
    if 'Wrong/Damaged Item' in top_issues: actions.append("Review packing quality — coordinate with warehouse")
    if 'Refund/Cancel'      in top_issues: actions.append("Process pending refunds in Seller Center")
    def store_summary(grp):
    platform, store_code, site_nick = grp.name

    return pd.Series({
        'PRIORITY':          priority,
        'PLATFORM': platform,
        'STORE_CODE': store_code,
        'SITE_NICK_NAME_ID': site_nick,
        # your metrics
        'TOTAL_CHATS':       len(grp),
        'COUNTRY':           country,
        'TOTAL_CHATS_7D':    len(grp),
        'UNREPLIED':         unreplied,
        'AUTO_REPLY_ONLY':   int(grp['ONLY_AUTO_REPLIED'].sum()),
        'UNRESOLVED':        unresolved,
        'ANGRY_BUYERS':      angry,
        'NEUTRAL_BUYERS':    int((grp['SENTIMENT']=='NEUTRAL').sum()),
        'HAPPY_BUYERS':      int((grp['SENTIMENT']=='HAPPY').sum()),
        'AVG_CSAT':          avg_csat,
        'TOP_ISSUES':        ' | '.join(top_issues),
        'ANGRY_BUYER_NAMES': ', '.join(angry_names) if angry_names else '-',
        'ACTION_ITEMS':      '\n'.join(f"• {a}" for a in actions) if actions else '✓ No critical actions',
        # safe optional columns
        'AVG_RESPONSE_TIME': grp['RESPONSE_TIME'].mean() if 'RESPONSE_TIME' in grp.columns else None,
        'TOTAL_ORDERS': grp['ORDER_ID'].nunique() if 'ORDER_ID' in grp.columns else None,
    })
   def build_excel(summary_df, trend_df, detail_df, today, seven_days_ago):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb  = openpyxl.Workbook()
    C_HDR_BG='FF1A3C6E'; C_HDR_FG='FFFFFFFF'; C_TITLE='FF0D2137'; C_ALT='FFF0F4FF'
    C_P1='FFFF4444'; C_P2='FFFF8C00'; C_P3='FFFFE000'; C_P4='FF00BB77'
    C_SHOPEE='FFEE4D2D'; C_LAZADA='FF0F146B'
    thin  = Side(style='thin', color='FFB0B8C4')
    thick = Side(style='medium', color='FF1A3C6E')

    def hdr(ws, row):
        for c in ws[row]:
            c.fill=PatternFill('solid',fgColor=C_HDR_BG)
            c.font=Font(bold=True,color=C_HDR_FG,size=10,name='Calibri')
            c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            c.border=Border(top=thick,bottom=thick,left=thin,right=thin)

    def banner(ws,title,sub,ncols):
        ws.row_dimensions[1].height=30; ws.row_dimensions[2].height=16
        for r,t,sz,b,fg,bg in [(1,title,14,True,'FFFFFFFF',C_TITLE),(2,sub,9,False,'FF888888','FFFAFAFA')]:
            c=ws.cell(row=r,column=1,value=t)
            c.font=Font(bold=b,size=sz,color=fg,name='Calibri')
            c.fill=PatternFill('solid',fgColor=bg)
            c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=ncols)

    def alt(ws,start,end,ncols):
        for r in range(start,end+1):
            if r%2==0:
                for c in range(1,ncols+1):
                    cell=ws.cell(row=r,column=c)
                    rgb=cell.fill.fgColor.rgb if cell.fill.patternType else '00000000'
                    if rgb in ('00000000','FFFFFFFF','FFFAFAFA'): cell.fill=PatternFill('solid',fgColor=C_ALT)

    def widths(ws,w):
        from openpyxl.utils import get_column_letter
        for i,v in enumerate(w,1): ws.column_dimensions[get_column_letter(i)].width=v

    PC={'P1 - CRITICAL':C_P1,'P2 - HIGH':C_P2,'P3 - MEDIUM':C_P3,'P4 - LOW':C_P4}
    PLC={'Shopee':C_SHOPEE,'Lazada':C_LAZADA}

    def platform_cell(c, val):
        col=PLC.get(str(val),'FFFAFAFA')
        c.fill=PatternFill('solid',fgColor=col)
        c.font=Font(bold=True,size=9,color='FFFFFFFF',name='Calibri')
        c.alignment=Alignment(horizontal='center',vertical='center')

    # Sheet 1
    ws1=wb.active; ws1.title='📊 Summary & Actions'
    S=[('PRIORITY','Priority'),('PLATFORM','Platform'),('STORE_CODE','Store Code'),
       ('SITE_NICK_NAME_ID','Account\n(Site Nick Name)'),('COUNTRY','Country'),
       ('TOTAL_CHATS_7D','Total Chats\n(7 Days)'),('UNREPLIED','Unreplied'),
       ('AUTO_REPLY_ONLY','Auto-Reply\nOnly'),('UNRESOLVED','Unresolved'),
       ('ANGRY_BUYERS','😡 Angry'),('NEUTRAL_BUYERS','😐 Neutral'),('HAPPY_BUYERS','😊 Happy'),
       ('AVG_CSAT','Avg CSAT\n(1-5)'),('TOP_ISSUES','Top Issues'),
       ('ANGRY_BUYER_NAMES','Angry Buyers'),('ACTION_ITEMS','Action Items')]
    banner(ws1,'🛒 Shopee + Lazada — Daily Summary & Action Items',
        f'Last 7 Days: {seven_days_ago} → {today}  |  Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}',len(S))
    HR=4; ws1.row_dimensions[HR].height=40
    for ci,(_,l) in enumerate(S,1): ws1.cell(row=HR,column=ci,value=l)
    hdr(ws1,HR)
    for ri,row in enumerate(summary_df.itertuples(),start=HR+1):
        ws1.row_dimensions[ri].height=64
        for ci,(f,_) in enumerate(S,1):
            val=getattr(row,f); c=ws1.cell(row=ri,column=ci,value=val)
            c.font=Font(size=9,name='Calibri'); c.alignment=Alignment(vertical='top',wrap_text=True)
            c.border=Border(left=thin,right=thin,top=thin,bottom=thin)
            if f=='PRIORITY':
                c.fill=PatternFill('solid',fgColor=PC.get(str(val)[:12],'FFFAFAFA'))
                c.font=Font(bold=True,size=9,color='FFFFFFFF' if 'P1' in str(val) or 'P2' in str(val) else '00000000',name='Calibri')
                c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            if f=='PLATFORM': platform_cell(c,val)
            if f=='ANGRY_BUYERS' and int(val or 0)>0:
                c.fill=PatternFill('solid',fgColor='FFFFCCCC'); c.font=Font(bold=True,size=9,color='FFCC0000',name='Calibri')
            if f=='AVG_CSAT':
                v=float(val or 0); c.fill=PatternFill('solid',fgColor='FFFFCCCC' if v<2 else('FFFFFF99' if v<3.5 else 'FFCCFFCC'))
                c.font=Font(bold=True,size=9,name='Calibri'); c.alignment=Alignment(horizontal='center',vertical='center')
    alt(ws1,HR+1,HR+len(summary_df),len(S))
    widths(ws1,[15,10,10,18,8,10,10,10,10,8,8,8,9,24,20,44]); ws1.freeze_panes=f'A{HR+1}'

    # Sheet 2
    ws2=wb.create_sheet('📅 7-Day Trend')
    T=[('PLATFORM','Platform'),('STORE_CODE','Store Code'),('SITE_NICK_NAME_ID','Account'),
       ('DATE','Date'),('TOTAL_CHATS','Total Chats'),('UNREPLIED','Unreplied'),
       ('UNRESOLVED','Unresolved'),('UNREPLIED_RATE_PCT','Unreplied\nRate %'),
       ('ANGRY','😡 Angry'),('NEUTRAL','😐 Neutral'),('HAPPY','😊 Happy'),
       ('AVG_CSAT','Avg CSAT'),('TOP_ISSUE','Top Issue')]
    banner(ws2,'📅 7-Day Trend — Shopee + Lazada',f'Date range: {seven_days_ago} → {today}',len(T))
    HR2=4; ws2.row_dimensions[HR2].height=40
    for ci,(_,l) in enumerate(T,1): ws2.cell(row=HR2,column=ci,value=l)
    hdr(ws2,HR2)
    for ri,row in enumerate(trend_df.itertuples(),start=HR2+1):
        ws2.row_dimensions[ri].height=18
        for ci,(f,_) in enumerate(T,1):
            val=getattr(row,f); c=ws2.cell(row=ri,column=ci,value=val)
            c.font=Font(size=9,name='Calibri'); c.alignment=Alignment(horizontal='center',vertical='center')
            c.border=Border(left=thin,right=thin,top=thin,bottom=thin)
            if f=='PLATFORM': platform_cell(c,val)
            if f=='ANGRY' and int(val or 0)>0: c.fill=PatternFill('solid',fgColor='FFFFCCCC');c.font=Font(bold=True,size=9,color='FFCC0000')
            if f=='HAPPY' and int(val or 0)>0: c.fill=PatternFill('solid',fgColor='FFCCFFCC')
            if f=='UNREPLIED' and int(val or 0)>0: c.fill=PatternFill('solid',fgColor='FFFFF4CC')
            if f=='AVG_CSAT':
                v=float(val or 0); c.fill=PatternFill('solid',fgColor='FFFFCCCC' if v<2 else('FFFFFF99' if v<3.5 else 'FFCCFFCC'))
                c.font=Font(bold=True,size=9)
    alt(ws2,HR2+1,HR2+len(trend_df),len(T))
    widths(ws2,[10,12,18,12,12,12,12,12,9,9,9,9,22]); ws2.freeze_panes=f'A{HR2+1}'

    # Sheet 3
    ws3=wb.create_sheet('🔍 Detailed Analysis')
    D=[('PLATFORM','Platform'),('STORE_CODE','Store Code'),('SITE_NICK_NAME_ID','Account\n(Site Nick Name)'),
       ('COUNTRY_CODE','Country'),('BUYER_NAME','Buyer Name'),
       ('LAST_MSG_TIME','Last Message'),
       ('IS_UNREPLIED','Unreplied?'),('ONLY_AUTO_REPLIED','Auto\nOnly?'),
       ('SELLER_STALLED','Stalled?\n(Will check)'),('IS_UNRESOLVED','Unresolved?'),
       ('SENTIMENT','Sentiment'),('CSAT_SCORE','CSAT\n(1-5)'),('ISSUE_CATEGORY','Issue Type'),
       ('BUYER_CHAT_SUMMARY','📋 Buyer Summary'),
       ('LAST_BUYER_MESSAGE','Last Buyer Message'),
       ('SUGGESTED_REPLY','💬 Suggested Reply'),
       ('ACTION_ITEM','⚡ Action Item'),
       ('CONVERSATION_ID','Conversation ID'),('FIRST_MSG_TIME','First Message'),
       ('AGE_HOURS','Age\n(hrs)'),('TOTAL_MESSAGES','Total\nMsgs'),
       ('BUYER_MESSAGES','Buyer\nMsgs'),('REAL_SELLER_REPLIES','Real\nReplies')]
    banner(ws3,'🔍 Detailed Chat Analysis — Last 7 Days (Shopee + Lazada)',
           'All conversations · Buyer Summary · Suggested Replies · ⚡ Action Items',len(D))
    HR3=4; ws3.row_dimensions[HR3].height=40
    for ci,(_,l) in enumerate(D,1): ws3.cell(row=HR3,column=ci,value=l)
    hdr(ws3,HR3)
    SF={'ANGRY':PatternFill('solid',fgColor='FFFFCCCC'),'HAPPY':PatternFill('solid',fgColor='FFCCFFCC'),'NEUTRAL':PatternFill('solid',fgColor='FFFFFF99')}
    for ri,row in enumerate(detail_df.itertuples(),start=HR3+1):
        ws3.row_dimensions[ri].height=55; sv=row.SENTIMENT
        for ci,(f,_) in enumerate(D,1):
            val=getattr(row,f,None); c=ws3.cell(row=ri,column=ci,value=val)
            c.font=Font(size=8,name='Calibri'); c.alignment=Alignment(vertical='top',wrap_text=True)
            c.border=Border(left=thin,right=thin,top=thin,bottom=thin)
            if f=='PLATFORM': platform_cell(c,val); c.font=Font(bold=True,size=8,color='FFFFFFFF',name='Calibri')
            if f=='SENTIMENT':
                c.fill=SF.get(sv,PatternFill()); c.font=Font(bold=True,size=8,color={'ANGRY':'FFCC0000','HAPPY':'FF006600','NEUTRAL':'FF666600'}.get(sv,''))
                c.alignment=Alignment(horizontal='center',vertical='center')
            if f=='CSAT_SCORE':
                v=int(val or 3); c.fill=PatternFill('solid',fgColor='FFFFCCCC' if v<=2 else('FFFFFF99' if v==3 else 'FFCCFFCC'))
                c.font=Font(bold=True,size=8); c.alignment=Alignment(horizontal='center',vertical='center')
            if f=='IS_UNREPLIED' and str(val)=='YES':
                c.fill=PatternFill('solid',fgColor='FFFFCC00'); c.font=Font(bold=True,size=8,color='FFCC6600'); c.alignment=Alignment(horizontal='center',vertical='center')
            if f=='SELLER_STALLED' and str(val)=='YES':
                c.fill=PatternFill('solid',fgColor='FFFFDDAA'); c.font=Font(bold=True,size=8,color='FF884400'); c.alignment=Alignment(horizontal='center',vertical='center')
            if f=='IS_UNRESOLVED' and str(val)=='YES':
                c.fill=PatternFill('solid',fgColor='FFFFEECC'); c.font=Font(bold=True,size=8,color='FFCC4400'); c.alignment=Alignment(horizontal='center',vertical='center')
            if f=='BUYER_CHAT_SUMMARY' and val:
                c.fill=PatternFill('solid',fgColor='FFE8F4FF'); c.font=Font(italic=True,size=8,color='FF003366',name='Calibri')
            if f=='ACTION_ITEM' and val:
                c.fill=PatternFill('solid',fgColor='FFFFF3CD'); c.font=Font(bold=True,size=8,color='FF7B4500',name='Calibri')
    alt(ws3,HR3+1,HR3+len(detail_df),len(D))
    widths(ws3,[10,10,18,8,18,14,10,8,10,10,10,7,14,32,28,38,34,20,14,7,8,8,8]); ws3.freeze_panes=f'A{HR3+1}'

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── FILE UPLOAD ────────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload your daily chat .xlsx file (must contain shopee_chat_enquiries and/or lazada_chat_enquiries sheets)",
    type=["xlsx"]
)

if uploaded:
    with st.spinner("Loading data from both platforms…"):
        xl = pd.ExcelFile(uploaded)
        frames = []
        for sheet in xl.sheet_names:
            if 'shopee' in sheet.lower() or 'lazada' in sheet.lower():
                df = xl.parse(sheet)
                df['PLATFORM'] = 'Shopee' if 'shopee' in sheet.lower() else 'Lazada'
                frames.append(df)
        if not frames:
            st.error("No Shopee or Lazada sheets found. Expected sheet names containing 'shopee' or 'lazada'.")
            st.stop()

        df = pd.concat(frames, ignore_index=True)
        df['MESSAGE_TIME'] = pd.to_datetime(df['MESSAGE_TIME'], errors='coerce')
        df = df.dropna(subset=['MESSAGE_TIME']).sort_values('MESSAGE_TIME')

    TODAY          = df['MESSAGE_TIME'].max().normalize()
    SEVEN_DAYS_AGO = TODAY - timedelta(days=6)

    shopee_rows = (df['PLATFORM']=='Shopee').sum()
    lazada_rows = (df['PLATFORM']=='Lazada').sum()
    st.success(f"✅ Loaded — **Shopee: {shopee_rows:,} rows** · **Lazada: {lazada_rows:,} rows** · Data up to **{TODAY.date()}**")

with st.spinner("Running analysis on both platforms…"):

    # ✅ Clean columns
    df.columns = df.columns.str.strip().str.upper()

    # ✅ Auto reply detection (FASTER)
    df['IS_AUTO_REPLY'] = (
        (df['SENDER'] == 'seller') &
        (df['MESSAGE_PARSED']
         .astype(str)
         .str.lower()
         .apply(lambda x: any(re.search(p, x) for p in AUTO_REPLY_PATTERNS)))
    )

    # ✅ Conversation level analysis
    conv_df = df.groupby('CONVERSATION_ID', group_keys=False).apply(analyse_conv).reset_index()

    # ✅ Last 7 days filter
    conv_7d = conv_df[conv_df['LAST_MSG_TIME'] >= SEVEN_DAYS_AGO].copy()
    conv_7d['DATE'] = conv_7d['LAST_MSG_TIME'].dt.date

    # ✅ Store summary
    summary_df = conv_7d.groupby(
        ['PLATFORM','STORE_CODE','SITE_NICK_NAME_ID']
    ).apply(store_summary).reset_index(drop=True)

    summary_df = summary_df.sort_values(['PRIORITY','PLATFORM','UNREPLIED'], ascending=[True,True,False])        df['IS_AUTO_REPLY'] = df.apply(
            lambda r: is_auto_reply(r['MESSAGE_PARSED']) if r['SENDER']=='seller' else False, axis=1
        )
        conv_7d.columns = (
    conv_7d.columns
    .str.strip()
    .str.upper()
)
       conv_df = df.groupby('CONVERSATION_ID', group_keys=False).apply(analyse_conv).reset_index()
        conv_7d = conv_df[conv_df['LAST_MSG_TIME'] >= SEVEN_DAYS_AGO].copy()
        conv_7d['DATE'] = conv_7d['LAST_MSG_TIME'].dt.date

        summary_df = conv_7d.groupby(['PLATFORM','STORE_CODE','SITE_NICK_NAME_ID'], group_keys=False).apply(store_summary).reset_index(drop=True)
        summary_df = summary_df.sort_values(['PRIORITY','PLATFORM','UNREPLIED'], ascending=[True,True,False])

        trend_df = conv_7d.groupby(['PLATFORM','STORE_CODE','SITE_NICK_NAME_ID','DATE']).agg(
            TOTAL_CHATS  =('CONVERSATION_ID','count'),
            UNREPLIED    =('IS_UNREPLIED','sum'),
            UNRESOLVED   =('IS_UNRESOLVED','sum'),
            ANGRY        =('SENTIMENT', lambda x:(x=='ANGRY').sum()),
            NEUTRAL      =('SENTIMENT', lambda x:(x=='NEUTRAL').sum()),
            HAPPY        =('SENTIMENT', lambda x:(x=='HAPPY').sum()),
            AVG_CSAT     =('CSAT_SCORE','mean'),
            TOP_ISSUE    =('ISSUE_CATEGORY', lambda x: x.mode().iloc[0] if len(x)>0 else '-'),
        ).reset_index()
        trend_df['UNREPLIED_RATE_PCT']=(trend_df['UNREPLIED']/trend_df['TOTAL_CHATS']*100).round(1)
        trend_df['AVG_CSAT']=trend_df['AVG_CSAT'].round(2)
        trend_df=trend_df.sort_values(['PLATFORM','STORE_CODE','DATE'])

        dcols=['PLATFORM','STORE_CODE','SITE_NICK_NAME_ID','COUNTRY_CODE','CONVERSATION_ID','BUYER_NAME',
               'FIRST_MSG_TIME','LAST_MSG_TIME','AGE_HOURS','TOTAL_MESSAGES','BUYER_MESSAGES',
               'SELLER_MESSAGES','REAL_SELLER_REPLIES','IS_UNREPLIED','ONLY_AUTO_REPLIED','SELLER_STALLED',
               'IS_UNRESOLVED','SENTIMENT','CSAT_SCORE','ISSUE_CATEGORY',
               'BUYER_CHAT_SUMMARY','LAST_BUYER_MESSAGE','SUGGESTED_REPLY','ACTION_ITEM']
        detail_df=conv_7d[dcols].copy()
        detail_df['IS_UNREPLIED']    =detail_df['IS_UNREPLIED'].map({True:'YES',False:'NO'})
        detail_df['ONLY_AUTO_REPLIED']=detail_df['ONLY_AUTO_REPLIED'].map({True:'YES',False:'NO'})
        detail_df['SELLER_STALLED']  =detail_df['SELLER_STALLED'].map({True:'YES',False:'NO'})
        detail_df['IS_UNRESOLVED']   =detail_df['IS_UNRESOLVED'].map({True:'YES',False:'NO'})
        detail_df=detail_df.sort_values(['PLATFORM','STORE_CODE','IS_UNREPLIED','SENTIMENT'],ascending=[True,True,False,True])

    # ── TODAY: build snapshot data ─────────────────────────────────────────────
    conv_today = conv_df[conv_df['LAST_MSG_TIME'].dt.normalize() == TODAY].copy()
    conv_today['DATE'] = conv_today['LAST_MSG_TIME'].dt.date

    today_summary_df = pd.DataFrame()
    today_detail_df  = pd.DataFrame()
    if len(conv_today) > 0:
        today_summary_df = conv_today.groupby(['PLATFORM','STORE_CODE','SITE_NICK_NAME_ID'], group_keys=False).apply(store_summary).reset_index(drop=True)
        today_summary_df = today_summary_df.sort_values(['PRIORITY','PLATFORM','UNREPLIED'], ascending=[True,True,False])
        td = conv_today[dcols].copy()
        td['IS_UNREPLIED']    = td['IS_UNREPLIED'].map({True:'YES',False:'NO'})
        td['ONLY_AUTO_REPLIED']= td['ONLY_AUTO_REPLIED'].map({True:'YES',False:'NO'})
        td['SELLER_STALLED']  = td['SELLER_STALLED'].map({True:'YES',False:'NO'})
        td['IS_UNRESOLVED']   = td['IS_UNRESOLVED'].map({True:'YES',False:'NO'})
        today_detail_df = td.sort_values(['PLATFORM','STORE_CODE','IS_UNREPLIED','SENTIMENT'], ascending=[True,True,False,True])

    # ── FILTERS (shared across both tabs) ─────────────────────────────────────
    st.markdown("---")
    st.subheader("🔍 Filters")
    col1, col2, col3 = st.columns(3)

    all_platforms = sorted(summary_df['PLATFORM'].dropna().unique().tolist())
    sel_platform  = col1.multiselect("Platform", all_platforms, default=all_platforms)

    filtered_by_platform = summary_df[summary_df['PLATFORM'].isin(sel_platform)] if sel_platform else summary_df
    all_stores = sorted(filtered_by_platform['STORE_CODE'].dropna().unique().tolist())
    sel_store  = col2.multiselect("Store Code", all_stores, default=all_stores)

    filtered_by_store = filtered_by_platform[filtered_by_platform['STORE_CODE'].isin(sel_store)] if sel_store else filtered_by_platform
    all_accounts = sorted(filtered_by_store['SITE_NICK_NAME_ID'].dropna().unique().tolist())
    sel_account  = col3.multiselect("Account (Site Nick Name ID)", all_accounts, default=all_accounts)

    def apply_filters(df):
        if df is None or len(df) == 0: return df
        mask = pd.Series([True]*len(df), index=df.index)
        if sel_platform: mask &= df['PLATFORM'].isin(sel_platform)
        if sel_store:    mask &= df['STORE_CODE'].isin(sel_store)
        if sel_account:  mask &= df['SITE_NICK_NAME_ID'].isin(sel_account)
        return df[mask]

    f_summary      = apply_filters(summary_df)
    f_trend        = apply_filters(trend_df)
    f_detail       = apply_filters(detail_df)
    f_today_sum    = apply_filters(today_summary_df) if len(today_summary_df)>0 else today_summary_df
    f_today_detail = apply_filters(today_detail_df)  if len(today_detail_df)>0  else today_detail_df

    def style_summary(df):
        def color_priority(val):
            colors = {'P1 - CRITICAL':'background-color:#ff4444;color:white',
                      'P2 - HIGH':'background-color:#ff8c00;color:white',
                      'P3 - MEDIUM':'background-color:#ffe000',
                      'P4 - LOW':'background-color:#00bb77;color:white'}
            return colors.get(val,'')
        def color_platform(val):
            return 'background-color:#ee4d2d;color:white' if val=='Shopee' else 'background-color:#0f146b;color:white'
        def color_angry(val):
            return 'background-color:#ffcccc;color:#cc0000;font-weight:bold' if int(val or 0)>0 else ''
        def color_csat(val):
            v=float(val or 0)
            if v<2: return 'background-color:#ffcccc'
            if v<3.5: return 'background-color:#ffff99'
            return 'background-color:#ccffcc'
        return df.style\
            .applymap(color_priority, subset=['PRIORITY'])\
            .applymap(color_platform, subset=['PLATFORM'])\
            .applymap(color_angry, subset=['ANGRY_BUYERS'])\
            .applymap(color_csat, subset=['AVG_CSAT'])

    def render_metrics(detail_df, summary_df, label):
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric(f"Conversations ({label})", f"{len(detail_df):,}")
        c2.metric("Accounts", len(summary_df))
        c3.metric("😡 Unreplied",   int((detail_df['IS_UNREPLIED']=='YES').sum())  if len(detail_df)>0 else 0)
        c4.metric("📭 Unresolved",  int((detail_df['IS_UNRESOLVED']=='YES').sum()) if len(detail_df)>0 else 0)
        c5.metric("😡 Angry Buyers", int(summary_df['ANGRY_BUYERS'].sum())          if len(summary_df)>0 else 0)
        c6.metric("Avg CSAT", f"{summary_df['AVG_CSAT'].mean():.2f} / 5"           if len(summary_df)>0 else "—")
        if len(summary_df)>0:
            p_counts = summary_df['PRIORITY'].value_counts()
            pcols = st.columns(4)
            for i,(p,icon) in enumerate([('P1 - CRITICAL','🔴'),('P2 - HIGH','🟠'),('P3 - MEDIUM','🟡'),('P4 - LOW','🟢')]):
                pcols[i].metric(f"{icon} {p}", p_counts.get(p,0))

    def render_summary_table(summary_df):
        if len(summary_df)==0:
            st.info("No data for this period / filter combination.")
            return
        disp = summary_df[['PRIORITY','PLATFORM','STORE_CODE','SITE_NICK_NAME_ID','COUNTRY',
                            'TOTAL_CHATS_7D','UNREPLIED','UNRESOLVED','ANGRY_BUYERS',
                            'NEUTRAL_BUYERS','HAPPY_BUYERS','AVG_CSAT','TOP_ISSUES','ACTION_ITEMS']]
        st.dataframe(style_summary(disp), use_container_width=True, height=420)

    def render_detail_table(detail_df):
        if len(detail_df)==0:
            st.info("No detailed conversations for this period / filter combination.")
            return
        show_cols = ['PLATFORM','STORE_CODE','SITE_NICK_NAME_ID','BUYER_NAME','IS_UNREPLIED',
                     'IS_UNRESOLVED','SENTIMENT','CSAT_SCORE','ISSUE_CATEGORY',
                     'BUYER_CHAT_SUMMARY','SUGGESTED_REPLY','ACTION_ITEM']
        avail = [c for c in show_cols if c in detail_df.columns]
        st.dataframe(detail_df[avail], use_container_width=True, height=450)

    # ── TABS ──────────────────────────────────────────────────────────────────
    st.markdown("---")
    tab_today, tab_7d = st.tabs([f"📅 Today  ({TODAY.date()})", f"📆 Last 7 Days  ({SEVEN_DAYS_AGO.date()} → {TODAY.date()})"])

    # ── TAB: TODAY ────────────────────────────────────────────────────────────
    with tab_today:
        if len(f_today_sum)==0:
            st.warning(f"No conversations found for today ({TODAY.date()}) after applying filters.")
        else:
            st.subheader(f"⚡ Today's Snapshot — {TODAY.date()}")
            render_metrics(f_today_detail, f_today_sum, "today")

            if len(sel_platform)>1:
                st.markdown("**Platform split**")
                pc2 = st.columns(2)
                for i,plat in enumerate(['Shopee','Lazada']):
                    pc = f_today_sum[f_today_sum['PLATFORM']==plat]
                    if len(pc)>0:
                        pc2[i].metric(f"{'🔴' if plat=='Shopee' else '🔵'} {plat}",
                            f"{pc['TOTAL_CHATS_7D'].sum():,} chats · {pc['UNREPLIED'].sum()} unreplied")

            st.subheader("📋 Today's Priority Summary")
            render_summary_table(f_today_sum)

            st.subheader("🔍 Today's Conversations")
            render_detail_table(f_today_detail)

            st.markdown("---")
            st.subheader("⬇️ Download Today's Report")
            # Build a single-day trend for today's download
            today_trend = conv_today.groupby(['PLATFORM','STORE_CODE','SITE_NICK_NAME_ID','DATE']).agg(
                TOTAL_CHATS=('CONVERSATION_ID','count'),
                UNREPLIED  =('IS_UNREPLIED','sum'),
                UNRESOLVED =('IS_UNRESOLVED','sum'),
                ANGRY      =('SENTIMENT',lambda x:(x=='ANGRY').sum()),
                NEUTRAL    =('SENTIMENT',lambda x:(x=='NEUTRAL').sum()),
                HAPPY      =('SENTIMENT',lambda x:(x=='HAPPY').sum()),
                AVG_CSAT   =('CSAT_SCORE','mean'),
                TOP_ISSUE  =('ISSUE_CATEGORY',lambda x:x.mode().iloc[0] if len(x)>0 else '-'),
            ).reset_index()
            today_trend['UNREPLIED_RATE_PCT'] = (today_trend['UNREPLIED']/today_trend['TOTAL_CHATS']*100).round(1)
            today_trend['AVG_CSAT'] = today_trend['AVG_CSAT'].round(2)
            f_today_trend = apply_filters(today_trend)
            with st.spinner("Building Excel…"):
                excel_today = build_excel(f_today_sum, f_today_trend, f_today_detail, TODAY.date(), TODAY.date())
            fname_today = f"Chat_Today_{TODAY.date()}_{'_'.join(sel_platform) if len(sel_platform)<3 else 'AllPlatforms'}.xlsx"
            st.download_button(
                label=f"📥 Download {fname_today}",
                data=excel_today, file_name=fname_today,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # ── TAB: 7-DAY VIEW ──────────────────────────────────────────────────────
    with tab_7d:
        st.subheader(f"📊 7-Day Overview — {SEVEN_DAYS_AGO.date()} → {TODAY.date()}")
        render_metrics(f_detail, f_summary, "7d")

        if len(sel_platform)>1:
            st.markdown("**Platform split**")
            pc3 = st.columns(2)
            for i,plat in enumerate(['Shopee','Lazada']):
                pc = f_summary[f_summary['PLATFORM']==plat]
                if len(pc)>0:
                    pc3[i].metric(f"{'🔴' if plat=='Shopee' else '🔵'} {plat}",
                        f"{pc['TOTAL_CHATS_7D'].sum():,} chats · {pc['UNREPLIED'].sum()} unreplied")

        st.subheader("📋 Summary & Action Items (7 Days)")
        render_summary_table(f_summary)

        st.subheader("🔍 All Conversations (7 Days)")
        render_detail_table(f_detail)

        st.markdown("---")
        st.subheader("⬇️ Download 7-Day Full Report")
        with st.spinner("Building formatted Excel…"):
            excel_bytes = build_excel(f_summary, f_trend, f_detail, TODAY.date(), SEVEN_DAYS_AGO.date())
        filename = f"Chat_7Day_{TODAY.date()}_{'_'.join(sel_platform) if len(sel_platform)<3 else 'AllPlatforms'}.xlsx"
        st.download_button(
            label=f"📥 Download {filename}",
            data=excel_bytes, file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.caption("3 sheets: 📊 Summary & Actions  ·  📅 7-Day Trend  ·  🔍 Detailed Analysis with suggested replies")
